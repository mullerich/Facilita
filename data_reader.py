"""
Esse programa tem o intuito de ler as tabelas de relatório disponibilizadas pelo INEP
https://sisu.mec.gov.br/#/relatorio#onepage, coletar os dados e transformar em lista
"""

from openpyxl import load_workbook

def open_excel(nome_arquivo, nome_folha):
    excel_data = load_workbook(nome_arquivo, read_only=True)
    datasheet = excel_data[nome_folha]     # Folha inscricao_2019_2
    return datasheet


def which_column(datasheet, colunas):
    """
    Determina o valor numérico de colunas de uma folha de planilha
    """
    num_colunas = datasheet.max_column
    colunas_usadas = []
    
    if colunas == []:
        colunas_usadas = [x for x in range(1, num_colunas+1)]

    else:
        for c in range(1, num_colunas+1):
            if datasheet.cell(1, c).value in colunas:
                colunas_usadas.append(c)
    
    return colunas_usadas


def leitor(datasheet, linha, colunas=[]):
    """
    Retorna a leitura de determinadas colunas de uma planilha em uma determinada linha
    """

    # Folha inscricao_2019_2
    
    dados = []
    colunas_usadas = which_column(datasheet, colunas)
    
    for col in colunas_usadas:
        dados.append(datasheet.cell(row=linha, column=col).value)
    
    return dados

